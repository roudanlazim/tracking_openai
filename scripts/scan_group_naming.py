import os
import json
import time
import pandas as pd
from tqdm import tqdm
from datetime import datetime
from dotenv import load_dotenv
from openai import OpenAI
from sentence_transformers import SentenceTransformer, util
import inspect
import argparse


# ===== CONFIG =====
excel_path = r"G:\Shared drives\Tech\PVR analysis\Tracking\PVR Scan Configuration\Scan group analysis\scan_group_analysis.xlsx"
column_name = "sg_name"
sheet_name = "sg_final"
similarity_threshold = 0.75
internal_similarity_threshold = 0.9
gpt_model = "gpt-4o"
embed_model = "all-MiniLM-L6-v2"
batch_size = 25
max_retries = 3
retry_backoff = 3.0
timestamp = datetime.now().strftime("%Y%m%d_%H%M")

from openpyxl import load_workbook

# Create an auto-incrementing sheet name
def get_unique_sheet_name(workbook_path, base_name="structured_analysis"):
    try:
        wb = load_workbook(workbook_path, read_only=True)
        existing_sheets = wb.sheetnames
        wb.close()
        versions = [
            int(name.replace(base_name + "_v", ""))
            for name in existing_sheets
            if name.startswith(base_name + "_v") and name.replace(base_name + "_v", "").isdigit()
        ]
        next_version = max(versions) + 1 if versions else 1
        return f"{base_name}_v{next_version}"
    except Exception as e:
        print(f"⚠️ Could not check workbook for sheet names: {e}")
        return f"{base_name}_v1"

# Update global sheet name
output_sheet = get_unique_sheet_name(excel_path)


# ===== PROGRESS LABELS =====
progress_labels = [
    "InTransit", "TransitHold", "CustomsHold", "Delivered", "Manifested", "DroppedOff", 
    "ShipmentCancelled", "Collected", "OutForDelivery", "DeliveryHold", "Returned", 
    "CollectionAttemptedButFailed", "DeliveryAttemptedButFailed", "BeingReturned", 
    "PartiallyDelivered", "DiscardedAbandonedLost"
]

# ===== Load OpenAI API Key =====
def load_api_key():
    dotenv_path = os.path.join(os.path.dirname(__file__), "..", "config", ".env")
    load_dotenv(dotenv_path=dotenv_path)
    return os.getenv("OPENAI_API_KEY")

client = OpenAI(api_key=load_api_key())

# ===== GPT Prompt Builder =====
def build_gpt_prompt(scan_batch: list[str]) -> list[dict]:
    progress_options = ", ".join(progress_labels)

    reference_examples = [
        {"Original": "Delivered to neighbour", "Structured": "Delivered - to neighbour"},
        {"Original": "Package damaged in transit", "Structured": "Damaged - in transit"},
        {"Original": "Collection failed - Parcel not ready", "Structured": "Collection failed - parcel not ready"},
        {"Original": "Delivery attempted - No access", "Structured": "Attempted delivery - no access"},
        {"Original": "Returned to sender", "Structured": "Returned - to sender"},
        {"Original": "Held by customs", "Structured": "Held - by customs"},
        {"Original": "Arrived at delivery depot", "Structured": "Arrival - at delivery depot"},
        {"Original": "Transit documentation missing. Clearance on hold", "Structured": "Held - missing documentation - clearance delay"},
        {"Original": "Delivery failed. Tax ID required", "Structured": "Failed delivery - tax ID required"},
        {"Original": "Package not collected - Damaged", "Structured": "Collection failed - damaged"},
        {"Original": "Warehouse Scan", "Structured": "Processed - scan"},
        {"Original": "Not collected", "Structured": "Collection failed - not collected"}
    ]

    naming_guide = (
        "Reformat each logistics scan group label using the following **modular naming structure**:\n\n"
        "**[Action] + [Detail or Modifier] + [Reason or Outcome] + [Instruction if needed]**\n\n"

        "**Core Rules:**\n"
        "1. Start with a clear **Action** verb: Delivered, Collected, Held, Attempted, Failed, Scheduled, Returned, Dropped, etc.\n"
        "2. Avoid vague phrasing like 'Not collected' or 'Shipment collected'. Instead, use:\n"
        "   - ✅ `Collection failed - not collected`\n"
        "   - ✅ `Collected`\n"
        "   - ✅ `Returned - to sender`\n"
        "3. If a label mentions any form of **scan**, use 'Processed', 'Arrival', or 'Dropped off' depending on context.\n"
        "4. If a label reflects a delay, include a **reason** such as 'due to weather', 'missing documentation', or 'inspection required'.\n"
        "5. If a label clearly maps to a progress state, use that progress label directly as the first term.\n"
        "6. Do NOT include filler words like 'package', 'shipment', or 'parcel' unless they disambiguate meaning.\n"

        "**Delivery Holds:**\n"
        "If a package is paused before delivery (e.g. held at depot or awaiting recipient action), use:\n"
        "→ `Delivery hold - at depot` or `Delivery hold - tax ID required`\n\n"

        "**Progress Labels (for progress_match):**\n"
        f"{progress_options}\n\n"

        "If a scan group cannot be clearly renamed, return it prefixed with **REVIEW:**.\n"
        "Your structured name must be **clear**, **modular**, **short**, and **semantically exact**.\n\n"

        "Output format (strict JSON):\n"
        '{ "results": [ {"structured_name": "...", "progress_match": "..." }, ... ] }'
    )

    examples = "\n".join([f"- {e['Original']} → {e['Structured']}" for e in reference_examples])
    scan_list = "\n".join([f"- {s}" for s in scan_batch])

    user_prompt = (
        f"{naming_guide}\n\n"
        f"Examples:\n{examples}\n\n"
        f"Now apply the same structure to the following scan group labels:\n{scan_list}\n\n"
        f"Return your response in valid JSON:\n"
        f'{{ "results": [ {{"structured_name": "...", "progress_match": "..." }}, ... ] }}\n\n'
        f"Use only the exact progress labels provided. If uncertain, use 'Uncertain'."
    )

    return [
        {"role": "system", "content": "You are an expert logistics classification assistant focused on AI clarity and modular naming."},
        {"role": "user", "content": user_prompt}
    ]

# ===== GPT Call =====
def call_gpt(messages):
    for attempt in range(1, max_retries + 1):
        try:
            response = client.chat.completions.create(
                model=gpt_model,
                messages=messages,
                temperature=0.2,
                response_format={"type": "json_object"}
            )
            content = response.choices[0].message.content.strip()
            return json.loads(content).get("results", [])
        except Exception as e:
            print(f"GPT failed on attempt {attempt}: {e}")
            if attempt < max_retries:
                time.sleep(retry_backoff * attempt)
    return []

# ===== Batch Logger =====
def log_gpt_batch(batch_num: int, scan_batch: list[str], gpt_results: list[dict]):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    script_name = os.path.splitext(os.path.basename(inspect.stack()[-1].filename))[0]
    base_log_dir = os.path.join(os.getcwd(), "logs", script_name)
    os.makedirs(base_log_dir, exist_ok=True)

    log_data = {
        "batch_number": batch_num,
        "timestamp": timestamp,
        "input_scan_groups": scan_batch,
        "gpt_results": gpt_results
    }

    file_name = f"gpt_batch_{str(batch_num).zfill(2)}_{timestamp}.json"
    log_path = os.path.join(base_log_dir, file_name)

    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(log_data, f, indent=2)

    print(f"📝 Logged GPT batch {batch_num} → {log_path}")

# ===== MAIN EXECUTION =====
def main(max_batches=None):
    print(f"🔹 Loading scan group data from sheet '{sheet_name}'...")
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    scan_groups = df[column_name].dropna().astype(str).tolist()

    print(f"🔸 Total scan group entries: {len(scan_groups)}")
    structured_names, gpt_progresses = [], []

    print("🚀 Calling GPT-4o for structured names and progress matches...")
    batch_counter = 0
    for i in tqdm(range(0, len(scan_groups), batch_size)):
        if max_batches is not None and batch_counter >= max_batches:
            break  # ✅ This line must be indented under the `if`

        batch = scan_groups[i:i+batch_size]
        messages = build_gpt_prompt(batch)
        results = call_gpt(messages)
        log_gpt_batch(batch_num=(i // batch_size) + 1, scan_batch=batch, gpt_results=results)

        if results:
            for item in results:
                structured_names.append(item.get("structured_name", "REVIEW"))
                gpt_progresses.append(item.get("progress_match", "Uncertain"))
        else:
            structured_names.extend([f"REVIEW: {s}" for s in batch])
            gpt_progresses.extend(["Uncertain"] * len(batch))

        batch_counter += 1

    scan_groups = scan_groups[:len(structured_names)]
    df = df.iloc[:len(structured_names)].copy()

    df["Structured Name (AI Ready)"] = structured_names
    df["Assigned Progress (GPT)"] = gpt_progresses

    print("🔹 Embedding structured names...")
    model = SentenceTransformer(embed_model)
    structured_embeddings = model.encode(structured_names, convert_to_tensor=True)
    progress_embeddings = model.encode(progress_labels, convert_to_tensor=True)

    seen = set()
    duplicates = []
    for name in scan_groups:
        if name in seen:
            duplicates.append("Yes")
        else:
            duplicates.append("No")
            seen.add(name)
    df["Exact Duplicate"] = duplicates

    similarity_flags = []
    for i, emb_i in enumerate(structured_embeddings):
        similar_count = 0
        for j, emb_j in enumerate(structured_embeddings):
            if i != j and float(util.cos_sim(emb_i, emb_j)[0]) >= internal_similarity_threshold:
                similar_count += 1
        similarity_flags.append("Yes" if similar_count > 0 else "No")
    df["Very Similar to Another Scan Group"] = similarity_flags

    print("🔹 Performing cosine similarity matching to progress states...")
    assigned_progress = []
    similarity_scores = []
    vague_flags = []
    match_consistency = []

    for i, emb in enumerate(structured_embeddings):
        scores = util.cos_sim(emb, progress_embeddings)[0]
        best_idx = scores.argmax()
        best_score = float(scores[best_idx])
        bm_label = progress_labels[best_idx] if best_score >= similarity_threshold else "Uncertain"
        assigned_progress.append(bm_label)
        similarity_scores.append(round(best_score, 3))
        vague_flags.append("Yes" if best_score < similarity_threshold else "")
        match_consistency.append("Yes" if bm_label == gpt_progresses[i] else "No")

    df["Assigned Progress (BM Match)"] = assigned_progress
    df["Progress Similarity Score"] = similarity_scores
    df["Vague or Overlapping (Score < 0.75)"] = vague_flags
    df["Progress Match Consistency"] = match_consistency

    print(f"✅ Writing results to sheet '{output_sheet}' in Excel...")
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=output_sheet, index=False)

    print("\nSummary:")
    print(f"Total scan groups: {len(scan_groups)}")
    print(f"Structured names generated: {len(structured_names)}")
    print(f"Vague progress entries (score < {similarity_threshold}): {vague_flags.count('Yes')}")
    print(f"Exact duplicates: {duplicates.count('Yes')}")
    print(f"Near-duplicates (semantic): {similarity_flags.count('Yes')}")
    print(f"GPT and BM progress matches aligned: {match_consistency.count('Yes')}")

if __name__ == "__main__":  # ✅ no leading space
    parser = argparse.ArgumentParser(description="Run scan group naming")
    parser.add_argument("--max-batches", type=int, default=None, help="Maximum number of GPT batches to run")
    args = parser.parse_args()
    main(max_batches=args.max_batches)
